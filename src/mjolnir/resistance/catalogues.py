"""Three catalogues, one normal form — and the seven traps that make it hard.

Every source here disagrees with the others about almost everything: what a
variant is called, whether a grade exists, which coordinate system applies, even
whether a row is about resistance at all. This module reduces all three to
:class:`CatalogueEntry` — one row per (catalogue, drug, variant) — so that
``consensus.py`` can compare like with like, and so that the report can print
which source said what.

The WHO loader is where the care goes, because the file is booby-trapped. Each
of the following was verified against the published files rather than inferred,
and each has code here that fails loudly rather than quietly:

1. **The header is on row 3.** Rows 1-2 are a merged banner. ``read_excel``
   with default arguments returns a table whose column names are fragments of a
   title.
2. **The repo's ``.txt`` master file is not the ``.xlsx``.** It carries 40,178
   rows against 48,152 and 14 drugs against 15 — Streptomycin is absent
   entirely, and rpoB, rpoC, rpsL and gid are affected. Mjolnir refuses it by
   name and by signature, and says why.
3. **The ``genomic position`` column is a decoy.** 38,884 of 48,152 rows carry
   the literal string ``(see "Genomic_coordinates" sheet)``. Coordinates are
   read from that sheet, never from this column.
4. **Grade strings are numeric-prefixed with a spaced ASCII hyphen** —
   ``2) Assoc w R - Interim``. The published PDF renders an en-dash; matching
   the PDF form matches nothing. ``config.normalise_grade`` owns the spellings.
5. **Grading is per (drug, variant).** ``inhA_c.-154G>A`` is Group 1 for
   isoniazid and Group 2 for ethionamide. The index is keyed on the pair, and
   the loader counts how many variants are graded differently across drugs so a
   test can see that nothing was collapsed.
6. **MNVs are decomposed and rejoined with ``&``.** One genomic change can map
   to several graded variants, so the coordinate table's ``variant`` field is
   split on the separator before indexing.
7. **Two columns are both literally named ``CHANGES vs ver1``.** The loader
   resolves columns positionally from a duplicate-aware index instead of
   assuming names are unique; a dict comprehension over the header silently
   discards one of them.

MTBseq and tbdb are simpler files with sharper edges of their own. MTBseq's list
is *flat*: it has no grading, so it can only ever contribute R or no-call, and
about a third of its rows are phylogenetic markers rather than resistance
mutations — an ``Antibiotic`` column reading ``phylo (EAI)`` is a lineage SNP
and loading it as a drug would invent resistance out of population structure.
tbdb carries WHO's grade strings without their numeric prefixes, so the mapping
is derived from ``config.WHO_GRADES`` rather than typed out a second time.

Nothing here decides a call. Grades become calls through ``config``'s published
mapping, and a row whose grading cannot be placed becomes ``no-call`` and is
counted — never a default of susceptible.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

from ..config import (
    CATALOGUE_MTBSEQ,
    CATALOGUE_TBDB,
    CATALOGUE_WHO,
    DRUG_ALIASES,
    H37RV_ACCESSION,
    MTBSEQ_ASYMMETRY_NOTE,
    SRC_MTBSEQ_MANUAL,
    SRC_TBDB,
    SRC_WHO_V2,
    WHO_COORDINATE_DECOY,
    WHO_GRADES,
    WHO_MNV_SEPARATOR,
    WHO_V2_EXPECTED_DRUGS,
    WHO_V2_EXPECTED_GENES,
    WHO_V2_EXPECTED_ROWS,
    WHO_V2_EXPECTED_VARIANTS,
    WHO_XLSX_COORDINATES_SHEET,
    WHO_XLSX_HEADER_ROW,
    WHO_XLSX_MASTER_SHEET,
    call_for_grade,
    normalise_drug,
    normalise_grade,
)
from ..records import (
    CALL_NO_CALL,
    CALL_R,
    CALL_R_INTERIM,
    CALL_R_OUTSIDE_WHO,
    CatalogueCall,
    DatabaseVersion,
    Variant,
)
from ..utils import LOG, MjolnirError, PathLike, natural_key, sha256sum, smart_open
from .normalise import (
    CoordinateKey,
    alias_keys,
    hgvs_key,
    is_rule_variant,
    normalise_hgvs,
    split_key,
)

# ---------------------------------------------------------------------------
# Licences and provenance, from design §12
# ---------------------------------------------------------------------------

LICENCE_WHO = "ODC-By v1.0 (redistributable with attribution)"
LICENCE_MTBSEQ = "GPL-3.0 (MTBseq)"
LICENCE_TBDB = "see the tbdb repository LICENCE at fetch time"

URL_WHO = "https://github.com/GTB-tbsequencing/mutation-catalogue-2023"
URL_MTBSEQ = "https://github.com/ngs-fzb/MTBseq_source"
URL_TBDB = "https://github.com/jodyphelan/tbdb"

#: Default file names under ``<db>/``, written by ``mjolnir db fetch``.
DEFAULT_WHO_XLSX = "who/WHO-UCN-TB-2023.7-eng.xlsx"
DEFAULT_WHO_COORDINATES = "who/WHO-UCN-TB-2023.7-eng_genomic_coordinates.txt"
DEFAULT_MTBSEQ_LIST = "mtbseq/MTB_Resistance_Mediating.txt"
DEFAULT_TBDB_CSV = "tbdb/mutations.csv"

FETCH_HINT = "mjolnir db fetch"

# ---------------------------------------------------------------------------
# The refusal
# ---------------------------------------------------------------------------

#: SOURCE: design §5.2, verified against the published files. The repo's plain
#: master file is a different and smaller object than the workbook, and reading
#: it would silently drop a drug and mis-grade four genes. This is a refusal, in
#: the same sense as ``config.kraken2_confidence`` refusing 0.0: there is no
#: flag that turns it into a warning, because a run that used it would produce a
#: report indistinguishable from a correct one.
WHO_TXT_REFUSAL = (
    "refusing to read the WHO catalogue from a plain-text master file.\n"
    "  The repo's WHO-UCN-TB-2023.x-eng_catalogue_master_file.txt is NOT "
    "equivalent to the .xlsx workbook: it carries 40,178 rows against the "
    "workbook's 48,152 and 14 drugs against 15 — Streptomycin is absent "
    "entirely — and the rpoB, rpoC, rpsL and gid grades differ. A run built on "
    "it would report no streptomycin result at all and would silently mis-grade "
    "four genes.\n"
    "  Use the workbook: Final Result Files/WHO-UCN-TB-2023.7-eng.xlsx from "
    + URL_WHO + ", or fetch it with: " + FETCH_HINT
)

#: The drug whose absence is the signature of a .txt-derived catalogue.
WHO_TXT_SIGNATURE_DRUG = "Streptomycin"

# ---------------------------------------------------------------------------
# WHO workbook layout, verified against
# WHO-UCN-TB-2023.6-eng_catalogue_master_file.txt (identical schema, 114 columns)
# and WHO-UCN-TB-2023.7-eng_genomic_coordinates.txt.
# ---------------------------------------------------------------------------

WHO_COL_DRUG = "drug"
WHO_COL_GENE = "gene"
WHO_COL_MUTATION = "mutation"
WHO_COL_VARIANT = "variant"
WHO_COL_TIER = "tier"
WHO_COL_EFFECT = "effect"
WHO_COL_POSITION = "genomic position"
WHO_COL_FINAL_GRADE = "FINAL CONFIDENCE GRADING"
WHO_COL_INITIAL_GRADE = "INITIAL CONFIDENCE GRADING"
WHO_COL_COMMENT = "Comment"
WHO_COL_ADDITIONAL = "Additional grading criteria applied"
WHO_COL_SILENT = "Silent mutation"
WHO_COL_FOOTNOTE = "Footnote"

#: Without these the workbook is not the WHO catalogue and the loader stops.
WHO_REQUIRED_COLUMNS: Tuple[str, ...] = (
    WHO_COL_DRUG, WHO_COL_GENE, WHO_COL_MUTATION, WHO_COL_VARIANT,
    WHO_COL_EFFECT, WHO_COL_FINAL_GRADE,
)
WHO_OPTIONAL_COLUMNS: Tuple[str, ...] = (
    WHO_COL_TIER, WHO_COL_POSITION, WHO_COL_INITIAL_GRADE, WHO_COL_COMMENT,
    WHO_COL_ADDITIONAL, WHO_COL_SILENT, WHO_COL_FOOTNOTE,
)

WHO_COORD_COLUMNS: Tuple[str, ...] = (
    "variant", "chromosome", "position", "reference_nucleotide",
    "alternative_nucleotide",
)

# ---------------------------------------------------------------------------
# tbdb, verified against jodyphelan/tbdb mutations.csv
# ---------------------------------------------------------------------------

TBDB_COLUMNS: Tuple[str, ...] = (
    "Gene", "Mutation", "type", "drug", "original_mutation", "confidence",
    "source", "comment",
)

#: tbdb writes WHO's grade strings with the numeric prefix stripped —
#: ``Not assoc w R - Interim`` rather than ``4) Not assoc w R - Interim``. The
#: mapping is derived from ``config.WHO_GRADES`` rather than retyped, so a
#: correction to the canonical spellings cannot leave this table behind.
TBDB_CONFIDENCE_TO_GRADE: Dict[str, str] = dict(
    (grade.split(")", 1)[1].strip().lower(), grade) for grade in WHO_GRADES
)

#: tbdb ``type`` values that assert a resistance association without supplying a
#: WHO grade. A row of one of these kinds with an unmappable confidence field is
#: read as R, which is exactly the §5.5 rule-3 path: surfaced as
#: ``R (outside WHO catalogue)`` downstream, never as a WHO Group 1 call.
#:
#: UNVERIFIED: every tbdb row sampled from the published mutations.csv carries
#: ``type=who_confidence``, so these two spellings are taken from TB-Profiler's
#: documented vocabulary and have not been observed in the file itself. The
#: consequence of the list being wrong is conservative in the right direction —
#: an unrecognised type yields ``no-call`` and is counted in ``skipped``, never
#: a susceptible result — and ``Catalogue.discrepancies`` prints the confidence
#: values a real load actually met.
TBDB_RESISTANCE_TYPES: Tuple[str, ...] = ("drug_resistance", "resistance_associated")

# ---------------------------------------------------------------------------
# MTBseq, verified against ngs-fzb/MTBseq_source var/res/MTB_Resistance_Mediating.txt
# ---------------------------------------------------------------------------

MTBSEQ_COL_POS_START = "Variant position genome start"
MTBSEQ_COL_TYPE = "Var. type"
MTBSEQ_COL_WT_BASE = "WT base"
MTBSEQ_COL_VAR_BASE = "Var. base"
MTBSEQ_COL_REGION = "Region"
MTBSEQ_COL_GENE_ID = "Gene ID"
MTBSEQ_COL_GENE_NAME = "Gene Name"
MTBSEQ_COL_AA_CHANGE = "AA change"
MTBSEQ_COL_GENE_POS = "Variant position gene start"
MTBSEQ_COL_ANTIBIOTIC = "Antibiotic"
MTBSEQ_COL_PMID = "Reference PMID"
MTBSEQ_COL_HIGH_CONF = "High Confidence SNP"
MTBSEQ_COL_COMMENT = "Comment"

MTBSEQ_REQUIRED_COLUMNS: Tuple[str, ...] = (
    MTBSEQ_COL_POS_START, MTBSEQ_COL_TYPE, MTBSEQ_COL_WT_BASE,
    MTBSEQ_COL_VAR_BASE, MTBSEQ_COL_REGION, MTBSEQ_COL_GENE_ID,
    MTBSEQ_COL_GENE_NAME, MTBSEQ_COL_AA_CHANGE, MTBSEQ_COL_ANTIBIOTIC,
)

#: ``Antibiotic`` values that are not antibiotics. About a third of the file is
#: phylogenetic marker SNPs — ``phylo (EAI Manila)``, ``phylo (Delhi/CAS, 3)`` —
#: which belong to lineage typing, not to resistance. Loading them as drug rows
#: would turn population structure into resistance calls.
MTBSEQ_NON_DRUG_PREFIXES: Tuple[str, ...] = ("phylo", "lineage", "-")

#: ``<word> (CODE)`` pairs, which is how MTBseq packs several drugs into one
#: cell: ``amikacin (AMK) kanamycin (KAN) capreomycin (CPR)``.
_MTBSEQ_DRUG_PAIR = re.compile(r"(?P<name>[A-Za-z][A-Za-z\-]*)\s*\((?P<code>[A-Za-z]+)\)")

#: MTBseq ``Region`` values and the HGVS prefix each implies.
_MTBSEQ_REGION_PREFIX = {"CDS": "c.", "PROM": "c.", "RNA": "n."}


# ---------------------------------------------------------------------------
# The normal form
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class CatalogueEntry:
    """One catalogue's statement about one variant for one drug.

    The unit is the **pair**, never the variant. ``inhA_c.-154G>A`` is Group 1
    for isoniazid and Group 2 for ethionamide, so a structure keyed on the
    variant name alone would have to pick one of those and would be wrong for
    whichever drug it discarded.

    ``grade`` is verbatim source text and is empty for MTBseq, which has no
    grading at all. ``rule_only`` marks the pooled names — ``katG_LoF``,
    ``pncA_deletion``, tbdb's ``frameshift`` — that describe a class of change
    and are matched by ``rules.py`` rather than by coordinate lookup.
    """

    catalogue: str
    drug: str
    gene: str
    hgvs: str
    call: str = CALL_NO_CALL
    grade: str = ""
    comment: str = ""
    effect: str = ""
    tier: Optional[int] = None
    evidence: str = ""
    #: Genomic coordinates that produce this graded variant, from the source's
    #: own coordinate table. Empty is normal for rule-only names and for every
    #: MTBseq indel, whose encoding is not VCF-normalised.
    coordinates: Tuple[CoordinateKey, ...] = ()
    rule_only: bool = False
    #: 1-based row in the source file, so a surprising grade can be looked up.
    source_row: int = 0

    @property
    def variant_key(self) -> str:
        return hgvs_key(self.gene, self.hgvs)

    @property
    def is_resistance(self) -> bool:
        return self.call in (CALL_R, CALL_R_INTERIM, CALL_R_OUTSIDE_WHO)

    def to_catalogue_call(self, catalogue_version: str = "",
                          catalogue_checksum: str = "",
                          matched_by: str = "") -> CatalogueCall:
        """Project into the record the report and the consensus engine read."""
        return CatalogueCall(
            catalogue=self.catalogue,
            drug=self.drug,
            grade=self.grade,
            comment=self.comment,
            source=_source_for_catalogue(self.catalogue),
            call=self.call,
            variant_key=self.variant_key,
            catalogue_version=catalogue_version,
            catalogue_checksum=catalogue_checksum,
            matched_by=matched_by,
            evidence=self.evidence,
        )


def _source_for_catalogue(catalogue: str) -> str:
    if catalogue == CATALOGUE_WHO:
        return SRC_WHO_V2
    if catalogue == CATALOGUE_MTBSEQ:
        return SRC_MTBSEQ_MANUAL
    if catalogue == CATALOGUE_TBDB:
        return SRC_TBDB
    return catalogue


@dataclass
class Catalogue:
    """One loaded catalogue, indexed both ways the design needs.

    ``by_coordinate`` serves WHO's own documented protocol — exact match on
    ``(chromosome, position, reference, alternative)`` — and ``by_key`` serves
    the cross-catalogue join, since MTBseq and tbdb have no coordinate table to
    match against. Aliases are folded into ``by_key`` at build time so that a
    lookup for ``rpoB_p.Ser450Leu`` finds a source that spelled it
    ``rpoB_p.Ser531Leu``, and the entry records which route matched.

    ``discrepancies`` holds everything that did not match the documented
    structure but was not fatal: an unexpected row count, an unrecognised grade,
    a drug nobody could name. They are printed with the catalogue's version in
    the report rather than logged and forgotten, because "the file was not quite
    what we expected" is exactly the sentence a reader needs before trusting a
    call.
    """

    name: str
    path: str = ""
    version: str = "unknown"
    checksum: str = ""
    licence: str = ""
    citation: str = ""
    url: str = ""
    entries: List[CatalogueEntry] = field(default_factory=list)
    discrepancies: List[str] = field(default_factory=list)
    note: str = ""
    #: Rows the loader read but could not place, by reason. Counted rather than
    #: dropped silently.
    skipped: Dict[str, int] = field(default_factory=dict)

    _by_key: Dict[str, List[CatalogueEntry]] = field(default_factory=dict, repr=False)
    _by_alias: Dict[str, List[CatalogueEntry]] = field(default_factory=dict, repr=False)
    _by_coordinate: Dict[CoordinateKey, List[CatalogueEntry]] = field(
        default_factory=dict, repr=False)

    # -- indexing ----------------------------------------------------------

    def index(self) -> "Catalogue":
        """(Re)build the lookup tables. Called once at the end of every loader."""
        self._by_key = {}
        self._by_alias = {}
        self._by_coordinate = {}
        for entry in self.entries:
            key = entry.variant_key
            if key:
                self._by_key.setdefault(key, []).append(entry)
                for alias in alias_keys(key):
                    self._by_alias.setdefault(alias, []).append(entry)
            for coordinate in entry.coordinates:
                self._by_coordinate.setdefault(coordinate, []).append(entry)
        return self

    # -- what is in it -----------------------------------------------------

    @property
    def drugs(self) -> Tuple[str, ...]:
        return tuple(sorted(set(e.drug for e in self.entries if e.drug), key=natural_key))

    @property
    def genes(self) -> Tuple[str, ...]:
        return tuple(sorted(set(e.gene for e in self.entries if e.gene), key=natural_key))

    @property
    def variant_keys(self) -> Tuple[str, ...]:
        return tuple(sorted(self._by_key or {}, key=natural_key))

    def multi_graded_variants(self) -> Dict[str, Dict[str, str]]:
        """Variants graded differently for different drugs.

        This is trap 5 made visible. ``inhA_c.-154G>A`` appears here with one
        grade for isoniazid and another for ethionamide, and its presence is
        what proves the loader did not deduplicate by variant name.
        """
        out: Dict[str, Dict[str, str]] = {}
        for key, entries in (self._by_key or {}).items():
            grades = dict((e.drug, e.grade) for e in entries if e.grade)
            if len(set(grades.values())) > 1:
                out[key] = grades
        return out

    # -- lookup ------------------------------------------------------------

    def lookup_coordinate(self, key: CoordinateKey) -> List[CatalogueEntry]:
        """WHO's primary protocol: exact match on chromosome, position, REF, ALT."""
        return list(self._by_coordinate.get(key, ()))

    def lookup_key(self, key: str, include_aliases: bool = True) -> List[CatalogueEntry]:
        """The cross-catalogue join, optionally through the numbering aliases."""
        canonical = _canonical_key(key)
        found = list(self._by_key.get(canonical, ()))
        if found or not include_aliases:
            return found
        return list(self._by_alias.get(canonical, ()))

    def entries_for(self, variant: Variant) -> List[Tuple[CatalogueEntry, str]]:
        """Every entry this catalogue has for an observed variant, with the route.

        Coordinate first, because that is WHO's documented protocol and it
        catches the alternative codon spellings a name lookup misses. The HGVS
        key is the fallback, and the alias table is the fallback to that — and
        the route is returned rather than discarded, so the annex can say
        whether a grade was matched on coordinates, on name, or only after
        translating a legacy codon number.
        """
        matched: List[Tuple[CatalogueEntry, str]] = []
        seen = set()
        for entry in self.lookup_coordinate(variant.coordinate_key):
            marker = id(entry)
            if marker not in seen:
                seen.add(marker)
                matched.append((entry, "coordinate"))
        key = _canonical_key(variant.hgvs_key)
        if key:
            for entry in self._by_key.get(key, ()):
                marker = id(entry)
                if marker not in seen:
                    seen.add(marker)
                    matched.append((entry, "hgvs"))
            for entry in self._by_alias.get(key, ()):
                marker = id(entry)
                if marker not in seen:
                    seen.add(marker)
                    matched.append((entry, "alias"))
        return matched

    def calls_for(self, variant: Variant) -> List[CatalogueCall]:
        """The :class:`records.CatalogueCall` rows for one observed variant."""
        return [
            entry.to_catalogue_call(self.version, self.checksum, route)
            for entry, route in self.entries_for(variant)
        ]

    # -- provenance --------------------------------------------------------

    def database_version(self) -> DatabaseVersion:
        note = self.note
        if self.discrepancies:
            note = "; ".join([note] + self.discrepancies) if note else \
                "; ".join(self.discrepancies)
        return DatabaseVersion(
            name=self.name, version=self.version, checksum=self.checksum,
            path=self.path, licence=self.licence,
            citation=_source_for_catalogue(self.name), url=self.url, note=note)

    def summary(self) -> Dict[str, Any]:
        return {
            "catalogue": self.name,
            "version": self.version,
            "checksum": self.checksum,
            "path": self.path,
            "entries": len(self.entries),
            "variants": len(self._by_key or {}),
            "coordinates": len(self._by_coordinate or {}),
            "drugs": list(self.drugs),
            "genes": len(self.genes),
            "rule_only_entries": sum(1 for e in self.entries if e.rule_only),
            "multi_graded_variants": len(self.multi_graded_variants()),
            "skipped": dict(self.skipped),
            "discrepancies": list(self.discrepancies),
        }

    def _skip(self, reason: str) -> None:
        self.skipped[reason] = self.skipped.get(reason, 0) + 1


def _canonical_key(key: str) -> str:
    """Normalise a join key without losing an already-canonical one."""
    text = str(key or "").strip()
    if not text:
        return ""
    gene, hgvs = split_key(text)
    if not gene:
        return text
    return hgvs_key(gene, normalise_hgvs(hgvs))


# ---------------------------------------------------------------------------
# Column resolution — trap 7
# ---------------------------------------------------------------------------

class _Header:
    """A duplicate-aware view of a header row.

    ``dict(zip(names, range(len(names))))`` is the obvious way to index a header
    and it is wrong for this file: the WHO master sheet has two columns both
    literally named ``CHANGES vs ver1``, and a dict keeps only the second. That
    is harmless for a column nobody reads and catastrophic for one somebody
    does, so the mapping is name -> list of positions and the duplicates are
    reported rather than resolved by luck.
    """

    def __init__(self, names: Sequence[Any]) -> None:
        self.names: List[str] = [_clean(n) for n in names]
        self.positions: Dict[str, List[int]] = {}
        for index, name in enumerate(self.names):
            if name:
                self.positions.setdefault(name.lower(), []).append(index)

    @property
    def duplicates(self) -> Dict[str, int]:
        """Repeated column names, spelled as the file spells them.

        The lookup index is case-folded, but the report has to quote the header
        back to a reader who will go looking for it in Excel, and
        ``changes vs ver1`` is not what they will see there.
        """
        return dict((self.names[where[0]], len(where))
                    for where in self.positions.values() if len(where) > 1)

    def has(self, name: str) -> bool:
        return name.lower() in self.positions

    def index_of(self, name: str) -> Optional[int]:
        where = self.positions.get(name.lower())
        return where[0] if where else None

    def value(self, row: Sequence[Any], name: str, default: str = "") -> str:
        index = self.index_of(name)
        if index is None or index >= len(row):
            return default
        return _clean(row[index])

    def missing(self, required: Iterable[str]) -> List[str]:
        return [name for name in required if not self.has(name)]


def _clean(value: Any) -> str:
    """A cell as text: never None, no stray whitespace, no float artefacts."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).split())


# ---------------------------------------------------------------------------
# WHO v2
# ---------------------------------------------------------------------------

def refuse_who_text_master(path: PathLike) -> None:
    """Raise if handed the repo's plain-text master file. Trap 2.

    Called by :func:`load_who` before anything is opened. There is deliberately
    no override: a run that read the ``.txt`` would produce a report with no
    streptomycin row and four silently mis-graded genes, and it would look
    exactly like a correct one.
    """
    name = Path(str(path)).name.lower()
    if name.endswith((".xlsx", ".xlsm")):
        return
    if "catalogue_master_file" in name or name.endswith((".txt", ".tsv")):
        raise MjolnirError("{0}\n  (offending file: {1})".format(WHO_TXT_REFUSAL, path))


def _who_version_from_name(path: Path) -> str:
    match = re.search(r"WHO-UCN-TB-([0-9]+\.[0-9]+)", path.name, re.IGNORECASE)
    if match:
        return "WHO-UCN-TB-{0}".format(match.group(1))
    return "unknown"


def _open_workbook(path: Path):
    """openpyxl, or a message naming the package to install.

    Imported here rather than at module scope so that ``mjolnir doctor`` can
    import this module and report the missing dependency instead of dying on it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise MjolnirError(
            "reading the WHO catalogue needs openpyxl, which is not installed.\n"
            "  pip install 'openpyxl>=3.0'   (or: conda install -c conda-forge openpyxl)\n"
            "  The catalogue is published only as .xlsx; the repo's .txt master "
            "file is not equivalent and Mjolnir will not read it."
        )
    try:
        return load_workbook(filename=str(path), read_only=True, data_only=True)
    except MjolnirError:
        raise
    except Exception as exc:  # openpyxl raises a zoo of exception types
        raise MjolnirError(
            "could not open {0} as an Excel workbook: {1}\n"
            "  Expected the WHO catalogue workbook with sheets {2!r} and {3!r}."
            .format(path, exc, WHO_XLSX_MASTER_SHEET, WHO_XLSX_COORDINATES_SHEET)
        )


def _sheet(workbook, wanted: str, path: Path):
    for name in workbook.sheetnames:
        if name.strip().lower() == wanted.strip().lower():
            return workbook[name]
    raise MjolnirError(
        "{0} has no sheet named {1!r}; it carries {2}.\n"
        "  This is not the WHO catalogue workbook. The published file is "
        "WHO-UCN-TB-2023.7-eng.xlsx from {3} — fetch it with: {4}".format(
            path, wanted, ", ".join(repr(n) for n in workbook.sheetnames),
            URL_WHO, FETCH_HINT)
    )


def _who_master_header(rows: Iterator[Sequence[Any]], path: Path) -> Tuple[_Header, int]:
    """Read past the merged banner and return the real header. Trap 1.

    The header is on row 3; rows 1-2 are a two-line merged title. The first two
    rows are read and *kept*, so that when row 3 turns out not to be a header
    the error can show what was actually there — a loader that only says
    "missing column 'drug'" sends the reader looking in the wrong place.
    """
    banner: List[str] = []
    header: Optional[_Header] = None
    read = 0
    for row in rows:
        read += 1
        if read < WHO_XLSX_HEADER_ROW:
            banner.append(" | ".join(_clean(cell) for cell in row[:4] if _clean(cell)))
            continue
        header = _Header(row)
        break
    if header is None:
        raise MjolnirError(
            "{0}: the {1!r} sheet has fewer than {2} rows, so it cannot carry the "
            "header on row {2}.".format(path, WHO_XLSX_MASTER_SHEET, WHO_XLSX_HEADER_ROW)
        )
    missing = header.missing(WHO_REQUIRED_COLUMNS)
    if missing:
        raise MjolnirError(
            "{0}: row {1} of {2!r} is not the catalogue header — it is missing "
            "{3}.\n"
            "  Row {1} is where the real header lives; rows 1-{4} are a merged "
            "banner, which read as: {5}\n"
            "  Row {1} read as: {6}\n"
            "  If this file has a different layout it is not the published WHO "
            "workbook; fetch it with: {7}".format(
                path, WHO_XLSX_HEADER_ROW, WHO_XLSX_MASTER_SHEET,
                ", ".join(repr(m) for m in missing), WHO_XLSX_HEADER_ROW - 1,
                " // ".join(b for b in banner if b) or "(empty)",
                ", ".join(repr(n) for n in header.names[:8]) or "(empty)",
                FETCH_HINT)
        )
    return header, read


def _split_graded_variants(text: str) -> List[str]:
    """Split a coordinate row's variant field on ``&``. Trap 6.

    One genomic change can decompose into several graded variants — that is how
    WHO encodes an MNV — and the VCF joins them with an ampersand in the
    ``graded_variant`` INFO field. Indexing the joined string would create a
    variant name no catalogue row contains, so the coordinate would match
    nothing and the MNV would be reported as ungraded.
    """
    raw = _clean(text)
    if not raw:
        return []
    return [part.strip() for part in raw.split(WHO_MNV_SEPARATOR) if part.strip()]


def _read_who_coordinates_sheet(sheet, path: Path) -> Dict[str, List[CoordinateKey]]:
    rows = sheet.iter_rows(values_only=True)
    header: Optional[_Header] = None
    # The coordinates sheet is not documented as carrying the master sheet's
    # merged banner, but a few leading rows are tolerated rather than assumed
    # away. The search is bounded so that a sheet with the wrong columns fails
    # in a moment instead of scanning two hundred thousand coordinate rows.
    for _ in range(WHO_XLSX_HEADER_ROW + 2):
        try:
            candidate = _Header(next(rows))
        except StopIteration:
            break
        if not candidate.missing(WHO_COORD_COLUMNS):
            header = candidate
            break
    if header is None:
        raise MjolnirError(
            "{0}: the {1!r} sheet has no header carrying {2}. Coordinates come "
            "from this sheet — the master sheet's `genomic position` column is "
            "a decoy and 38,884 of its rows hold the literal string {3!r}."
            .format(path, WHO_XLSX_COORDINATES_SHEET,
                    ", ".join(WHO_COORD_COLUMNS), WHO_COORDINATE_DECOY)
        )
    return _collect_coordinates(
        ((header.value(row, "variant"), header.value(row, "chromosome"),
          header.value(row, "position"), header.value(row, "reference_nucleotide"),
          header.value(row, "alternative_nucleotide")) for row in rows),
        path)


def _collect_coordinates(rows: Iterable[Tuple[str, str, str, str, str]],
                         path: Path) -> Dict[str, List[CoordinateKey]]:
    """Build ``graded variant -> coordinates`` from five-column rows."""
    out: Dict[str, List[CoordinateKey]] = {}
    for variant, chrom, position, ref, alt in rows:
        if not variant or not position:
            continue
        try:
            pos = int(float(position))
        except (TypeError, ValueError):
            raise MjolnirError(
                "{0}: coordinate row for {1!r} has a non-numeric position {2!r}. "
                "Coordinates must come from the {3} table, not from the master "
                "sheet's decoy `genomic position` column.".format(
                    path, variant, position, WHO_XLSX_COORDINATES_SHEET)
            )
        key = (chrom or H37RV_ACCESSION, pos, ref.upper(), alt.upper())
        for name in _split_graded_variants(variant):
            out.setdefault(name, []).append(key)
    return out


def read_who_coordinates_file(path: PathLike) -> Dict[str, List[CoordinateKey]]:
    """Coordinates from WHO's standalone table, as ``.txt``/``.tsv`` or VCF.

    Both published forms are accepted because both are shipped:
    ``WHO-UCN-TB-2023.7-eng_genomic_coordinates.txt`` (five named columns) and
    ``Genomic_coordinates_7May2024.vcf.gz`` (``graded_variant=`` in INFO, with
    MNVs joined by ``&``). This is the *only* sanctioned source of coordinates —
    the master sheet's own position column is the decoy of trap 3.
    """
    resolved = Path(str(path)).expanduser()
    if not resolved.exists():
        raise MjolnirError(
            "WHO genomic coordinates not found at {0}.\n  fetch them with: {1}"
            .format(resolved, FETCH_HINT))
    name = resolved.name.lower()
    if ".vcf" in name:
        return _collect_coordinates(_iter_vcf_coordinates(resolved), resolved)
    with smart_open(resolved, "rt") as handle:
        reader = csv.reader(handle, delimiter="\t")
        try:
            header = _Header(next(reader))
        except StopIteration:
            raise MjolnirError("{0} is empty".format(resolved))
        missing = header.missing(WHO_COORD_COLUMNS)
        if missing:
            raise MjolnirError(
                "{0} is missing the coordinate column(s) {1}; expected {2}"
                .format(resolved, ", ".join(missing), ", ".join(WHO_COORD_COLUMNS)))
        return _collect_coordinates(
            ((header.value(row, "variant"), header.value(row, "chromosome"),
              header.value(row, "position"), header.value(row, "reference_nucleotide"),
              header.value(row, "alternative_nucleotide")) for row in reader),
            resolved)


def _iter_vcf_coordinates(path: Path) -> Iterator[Tuple[str, str, str, str, str]]:
    with smart_open(path, "rt") as handle:
        for line in handle:
            if line.startswith("#"):
                continue
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 8:
                continue
            info = fields[7]
            graded = ""
            for item in info.split(";"):
                if item.startswith("graded_variant="):
                    graded = item.split("=", 1)[1]
                    break
            if not graded:
                continue
            yield (graded, fields[0], fields[1], fields[3], fields[4])


def load_who(path: PathLike, coordinates_path: Optional[PathLike] = None,
             strict: bool = True) -> Catalogue:
    """Load the WHO v2 workbook into the normal form.

    ``strict`` governs *counts*, not structure. A missing sheet, a header that
    is not on row 3 or an absent required column is always fatal, because the
    file is then not the catalogue. Row and drug counts that differ from the
    published v2 figures are fatal under ``strict`` and become recorded
    discrepancies without it — a 3rd edition was called for in 2024 and this
    loader must not be what refuses to read it, while a v2 that quietly lost
    8,000 rows must not pass unnoticed either.
    """
    refuse_who_text_master(path)
    resolved = Path(str(path)).expanduser()
    if not resolved.exists():
        raise MjolnirError(
            "WHO catalogue workbook not found at {0}.\n"
            "  fetch it with: {1}\n"
            "  (source: {2}, Final Result Files/WHO-UCN-TB-2023.7-eng.xlsx)"
            .format(resolved, FETCH_HINT, URL_WHO))

    catalogue = Catalogue(
        name=CATALOGUE_WHO, path=str(resolved), version=_who_version_from_name(resolved),
        checksum=sha256sum(resolved), licence=LICENCE_WHO, citation=SRC_WHO_V2,
        url=URL_WHO,
        note="grades are per (drug, variant); coordinates come from the {0} table, "
             "never from the master sheet's `genomic position` column"
             .format(WHO_XLSX_COORDINATES_SHEET))

    workbook = _open_workbook(resolved)
    try:
        master = _sheet(workbook, WHO_XLSX_MASTER_SHEET, resolved)
        if coordinates_path is not None:
            coordinates = read_who_coordinates_file(coordinates_path)
            catalogue.note += "; coordinates read from {0}".format(coordinates_path)
        else:
            coordinates = _read_who_coordinates_sheet(
                _sheet(workbook, WHO_XLSX_COORDINATES_SHEET, resolved), resolved)

        rows = master.iter_rows(values_only=True)
        header, header_row = _who_master_header(rows, resolved)
        if header.duplicates:
            # Trap 7. Reported, not repaired: the duplicates in the published
            # file are both `CHANGES vs ver1` and Mjolnir reads neither, but a
            # duplicate over a column it *does* read must be visible.
            catalogue.discrepancies.append(
                "duplicate column names on the master sheet: {0}".format(
                    ", ".join("{0} x{1}".format(n, c)
                              for n, c in sorted(header.duplicates.items()))))

        decoy_rows = 0
        real_position_rows = 0
        unknown_grades: Dict[str, int] = {}
        pairs = set()
        row_number = header_row
        for row in rows:
            row_number += 1
            drug_raw = header.value(row, WHO_COL_DRUG)
            variant_raw = header.value(row, WHO_COL_VARIANT)
            if not drug_raw and not variant_raw:
                continue
            if not drug_raw or not variant_raw:
                catalogue._skip("row missing drug or variant")
                continue

            position_cell = header.value(row, WHO_COL_POSITION)
            if position_cell:
                if position_cell == WHO_COORDINATE_DECOY or position_cell.startswith("(see"):
                    decoy_rows += 1
                else:
                    real_position_rows += 1

            gene = header.value(row, WHO_COL_GENE)
            mutation = header.value(row, WHO_COL_MUTATION)
            gene, hgvs = _who_variant_parts(gene, mutation, variant_raw)

            grade_raw = header.value(row, WHO_COL_FINAL_GRADE) or \
                header.value(row, WHO_COL_INITIAL_GRADE)
            grade = normalise_grade(grade_raw)
            if grade_raw and not grade:
                unknown_grades[grade_raw] = unknown_grades.get(grade_raw, 0) + 1

            tier_text = header.value(row, WHO_COL_TIER)
            try:
                tier = int(tier_text) if tier_text else None
            except ValueError:
                tier = None

            comment = header.value(row, WHO_COL_COMMENT)
            additional = header.value(row, WHO_COL_ADDITIONAL)
            footnote = header.value(row, WHO_COL_FOOTNOTE)
            evidence = "; ".join(part for part in (
                "additional grading criteria: " + additional if additional else "",
                "footnote: " + footnote if footnote else "",
            ) if part)

            drug = normalise_drug(drug_raw)
            entry = CatalogueEntry(
                catalogue=CATALOGUE_WHO,
                drug=drug,
                gene=gene,
                hgvs=hgvs,
                call=call_for_grade(grade) if grade else CALL_NO_CALL,
                grade=grade or grade_raw,
                comment=comment,
                effect=header.value(row, WHO_COL_EFFECT),
                tier=tier,
                evidence=evidence,
                coordinates=tuple(coordinates.get(variant_raw, ())),
                rule_only=is_rule_variant(hgvs),
                source_row=row_number,
            )
            pairs.add((entry.drug, entry.variant_key))
            catalogue.entries.append(entry)
    finally:
        # read_only workbooks hold an open zip handle; closing it can only fail
        # for filesystem reasons, and by this point every row is already in
        # memory, so the failure is logged rather than raised over the result.
        try:
            workbook.close()
        except OSError as exc:
            LOG.debug("closing %s failed after reading it: %s", resolved, exc)

    catalogue.index()
    _check_who_shape(catalogue, decoy_rows, real_position_rows, unknown_grades,
                     len(pairs), strict)
    LOG.info("WHO catalogue: %d graded (drug, variant) rows, %d unique variants, "
             "%d drugs, %d genes", len(catalogue.entries), len(catalogue.variant_keys),
             len(catalogue.drugs), len(catalogue.genes))
    return catalogue


def _who_variant_parts(gene: str, mutation: str, variant: str) -> Tuple[str, str]:
    """The (gene, hgvs) pair for a master-sheet row.

    The sheet supplies all three of ``gene``, ``mutation`` and ``variant``, and
    ``variant`` is documented as their concatenation (``bacA`` + ``c.102G>A`` =
    ``bacA_c.102G>A``). Preferring the separate columns and falling back to
    splitting ``variant`` keeps the loader working on a row where one of them is
    blank, which does happen for the pooled LoF names.
    """
    if gene and mutation:
        return gene, normalise_hgvs(mutation)
    split_gene, split_hgvs = split_key(variant)
    return (gene or split_gene), normalise_hgvs(mutation or split_hgvs)


def _check_who_shape(catalogue: Catalogue, decoy_rows: int, real_position_rows: int,
                     unknown_grades: Dict[str, int], pair_count: int,
                     strict: bool) -> None:
    """Compare the loaded workbook against the published v2 figures.

    Structure has already been enforced; this is about size and coverage, and
    its most important job is trap 2's second line of defence. A workbook built
    from the plain-text master file has no Streptomycin rows at all, and that
    absence is a signature no row count can explain away.
    """
    problems: List[str] = []
    drugs = catalogue.drugs

    if WHO_TXT_SIGNATURE_DRUG not in drugs:
        message = (
            "the WHO catalogue at {0} grades {1} drugs and {2} is not among "
            "them: {3}.\n"
            "  That absence is the signature of the repo's plain-text master "
            "file rather than the published workbook.\n{4}".format(
                catalogue.path, len(drugs), WHO_TXT_SIGNATURE_DRUG,
                ", ".join(drugs), WHO_TXT_REFUSAL)
        )
        if strict:
            raise MjolnirError(message)
        # strict=False is the deliberate "this is a newer edition" override, so
        # it demotes even this check — but the reason is carried into the
        # report rather than dropped, because a run with no streptomycin rows
        # must never look like a run that found streptomycin susceptible.
        catalogue.discrepancies.append(message)
        LOG.warning("%s", message)

    if len(catalogue.entries) != WHO_V2_EXPECTED_ROWS:
        problems.append(
            "{0} (drug, variant) rows, expected {1} for WHO v2".format(
                len(catalogue.entries), WHO_V2_EXPECTED_ROWS))
    if len(catalogue.variant_keys) != WHO_V2_EXPECTED_VARIANTS:
        problems.append(
            "{0} unique variants, expected {1} for WHO v2".format(
                len(catalogue.variant_keys), WHO_V2_EXPECTED_VARIANTS))
    if len(drugs) != WHO_V2_EXPECTED_DRUGS:
        problems.append("{0} drugs, expected {1} for WHO v2".format(
            len(drugs), WHO_V2_EXPECTED_DRUGS))
    if len(catalogue.genes) != WHO_V2_EXPECTED_GENES:
        problems.append("{0} genes, expected {1} for WHO v2".format(
            len(catalogue.genes), WHO_V2_EXPECTED_GENES))
    if pair_count != len(catalogue.entries):
        problems.append(
            "{0} rows collapse to {1} distinct (drug, variant) pairs; the "
            "workbook is expected to carry one row per pair".format(
                len(catalogue.entries), pair_count))
    if not catalogue._by_coordinate:
        problems.append(
            "no genomic coordinates were loaded, so WHO's documented "
            "coordinate-based matching protocol cannot be applied")
    if decoy_rows == 0 and real_position_rows == 0:
        problems.append(
            "the master sheet carried no `genomic position` column at all")
    for text, count in sorted(unknown_grades.items()):
        problems.append(
            "{0} row(s) carry the grading {1!r}, which is not one of WHO's five "
            "grade strings; those rows produce no call".format(count, text))

    if not problems:
        return
    detail = "; ".join(problems)
    if strict:
        raise MjolnirError(
            "the WHO catalogue at {0} does not match the published v2 figures: "
            "{1}.\n"
            "  If this is a newer edition, load it with strict=False "
            "(--allow-catalogue-drift) and the differences will be printed in "
            "the report instead.".format(catalogue.path, detail)
        )
    catalogue.discrepancies.append(detail)
    LOG.warning("WHO catalogue differs from the published v2 figures: %s", detail)


# ---------------------------------------------------------------------------
# MTBseq
# ---------------------------------------------------------------------------

def _parse_mtbseq_drugs(cell: str) -> Tuple[List[str], List[str]]:
    """Drugs out of ``amikacin (AMK) kanamycin (KAN) capreomycin (CPR)``.

    Both halves of each pair are tried against ``config.normalise_drug`` because
    neither alone is reliable: MTBseq writes ``capreomycin (CPR)`` where the
    catalogue's own code is ``CAP``. Anything that resolves to nothing is
    returned as unresolved rather than passed through, since ``normalise_drug``
    echoes unknown names and an echoed fragment would become a phantom drug
    column in the report.
    """
    text = _clean(cell)
    if not text:
        return [], []
    known: List[str] = []
    unresolved: List[str] = []
    pairs = _MTBSEQ_DRUG_PAIR.findall(text)
    candidates: List[Tuple[str, str]] = list(pairs)
    if not candidates:
        candidates = [(token, "") for token in text.split()]
    for name, code in candidates:
        for attempt in (name, code):
            # DRUG_ALIASES rather than normalise_drug(): the latter echoes an
            # unknown name back, which would turn "mediating" or a stray word
            # into a drug column in the report.
            resolved = DRUG_ALIASES.get(attempt.lower()) if attempt else None
            if resolved:
                if resolved not in known:
                    known.append(resolved)
                break
        else:
            if name and name not in unresolved:
                unresolved.append(name)
    return known, unresolved


def _mtbseq_hgvs(region: str, aa_change: str, wt_base: str, var_base: str,
                 position: str) -> str:
    """Build an HGVS name from MTBseq's own columns.

    MTBseq writes three different things in ``AA change`` depending on
    ``Region``: a protein change for CDS rows (``Arg738Gln``), a promoter
    nucleotide change for PROM rows (``-14c>t``), and an rRNA position for RNA
    rows (``1401a>g``). The prefix follows the region — ``n.`` for rRNA — which
    matters because ``rrs_n.1401A>G`` and ``rrs_c.1401A>G`` do not join.
    """
    change = _clean(aa_change)
    prefix = _MTBSEQ_REGION_PREFIX.get(_clean(region).upper(), "c.")
    if change and change != "-":
        return normalise_hgvs(change, default_prefix=prefix)
    # No named change: fall back to the gene-relative position and the alleles.
    pos = _clean(position)
    if pos and wt_base and var_base:
        return normalise_hgvs("{0}{1}>{2}".format(pos, wt_base, var_base),
                              default_prefix=prefix)
    return ""


def load_mtbseq(path: PathLike, extended: Optional[PathLike] = None) -> Catalogue:
    """Load MTBseq's flat resistance list.

    Flat means exactly what it says: there is no confidence grading anywhere in
    the file, so every retained row can only contribute ``R`` or nothing. That
    asymmetry travels with the catalogue in ``note`` and is stated in the report
    rather than papered over with a manufactured grade — an MTBseq row agreeing
    with a WHO Group 1 call is weaker corroboration than it looks.

    Roughly a third of the file is phylogenetic marker SNPs whose ``Antibiotic``
    column reads ``phylo (EAI)``. Those are lineage-defining positions; loading
    them as drug rows would convert population structure into resistance.
    """
    resolved = Path(str(path)).expanduser()
    if not resolved.exists():
        raise MjolnirError(
            "MTBseq resistance list not found at {0}.\n"
            "  fetch it with: {1}\n"
            "  (source: {2}, var/res/MTB_Resistance_Mediating.txt)"
            .format(resolved, FETCH_HINT, URL_MTBSEQ))

    catalogue = Catalogue(
        name=CATALOGUE_MTBSEQ, path=str(resolved), checksum=sha256sum(resolved),
        version="MTBseq var/res list", licence=LICENCE_MTBSEQ,
        citation=SRC_MTBSEQ_MANUAL, url=URL_MTBSEQ, note=MTBSEQ_ASYMMETRY_NOTE)

    sources = [resolved]
    if extended is not None:
        extended_path = Path(str(extended)).expanduser()
        if extended_path.exists():
            sources.append(extended_path)
        else:
            catalogue.discrepancies.append(
                "extended list {0} was requested but is not present".format(extended_path))

    unresolved_drugs: Dict[str, int] = {}
    for source in sources:
        _load_mtbseq_file(source, catalogue, unresolved_drugs)

    if unresolved_drugs:
        catalogue.discrepancies.append(
            "unrecognised antibiotic name(s) in {0}: {1}".format(
                resolved.name,
                ", ".join("{0} x{1}".format(name, count)
                          for name, count in sorted(unresolved_drugs.items()))))
    catalogue.index()
    LOG.info("MTBseq list: %d (drug, variant) rows over %d variants, %d "
             "phylogenetic-marker rows skipped", len(catalogue.entries),
             len(catalogue.variant_keys),
             catalogue.skipped.get("phylogenetic marker, not a drug", 0))
    return catalogue


def _load_mtbseq_file(path: Path, catalogue: Catalogue,
                      unresolved_drugs: Dict[str, int]) -> None:
    with smart_open(path, "rt") as handle:
        reader = csv.reader(handle, delimiter="\t")
        try:
            first = next(reader)
        except StopIteration:
            raise MjolnirError("{0} is empty".format(path))
        if first and first[0].startswith("#"):
            first = [first[0].lstrip("#")] + list(first[1:])
        header = _Header(first)
        missing = header.missing(MTBSEQ_REQUIRED_COLUMNS)
        if missing:
            raise MjolnirError(
                "{0} is missing the column(s) {1}.\n"
                "  Expected MTBseq's 25-column var/res table, whose header line "
                "begins '#Variant position genome start'. If MTBseq changed the "
                "format, re-fetch with: {2}".format(
                    path, ", ".join(repr(m) for m in missing), FETCH_HINT))

        row_number = 1
        for row in reader:
            row_number += 1
            if not row or not any(_clean(cell) for cell in row):
                continue
            antibiotic = header.value(row, MTBSEQ_COL_ANTIBIOTIC)
            lowered = antibiotic.lower()
            if not antibiotic or any(lowered.startswith(p)
                                     for p in MTBSEQ_NON_DRUG_PREFIXES):
                catalogue._skip("phylogenetic marker, not a drug")
                continue

            drugs, unresolved = _parse_mtbseq_drugs(antibiotic)
            for name in unresolved:
                unresolved_drugs[name] = unresolved_drugs.get(name, 0) + 1
            if not drugs:
                catalogue._skip("no recognised drug in the Antibiotic column")
                continue

            gene = header.value(row, MTBSEQ_COL_GENE_NAME)
            if not gene or gene == "-":
                gene = header.value(row, MTBSEQ_COL_GENE_ID)
            region = header.value(row, MTBSEQ_COL_REGION)
            hgvs = _mtbseq_hgvs(
                region,
                header.value(row, MTBSEQ_COL_AA_CHANGE),
                header.value(row, MTBSEQ_COL_WT_BASE),
                header.value(row, MTBSEQ_COL_VAR_BASE),
                header.value(row, MTBSEQ_COL_GENE_POS))
            if not gene or not hgvs:
                catalogue._skip("row has no usable gene or variant name")
                continue

            coordinates = _mtbseq_coordinates(header, row, catalogue)
            comment = header.value(row, MTBSEQ_COL_COMMENT)
            evidence = "; ".join(part for part in (
                "PMID/reference: " + header.value(row, MTBSEQ_COL_PMID)
                if header.value(row, MTBSEQ_COL_PMID) not in ("", "-") else "",
                "MTBseq 'High Confidence SNP' column: "
                + header.value(row, MTBSEQ_COL_HIGH_CONF)
                if header.value(row, MTBSEQ_COL_HIGH_CONF) not in ("", "-") else "",
            ) if part)

            for drug in drugs:
                catalogue.entries.append(CatalogueEntry(
                    catalogue=CATALOGUE_MTBSEQ,
                    drug=drug,
                    gene=gene,
                    hgvs=hgvs,
                    # Flat list: R or nothing. There is no grade to record.
                    call=CALL_R,
                    grade="",
                    comment=comment if comment != "-" else "",
                    effect=region,
                    evidence=evidence,
                    coordinates=coordinates,
                    rule_only=is_rule_variant(hgvs),
                    source_row=row_number,
                ))


def _mtbseq_coordinates(header: _Header, row: Sequence[Any],
                        catalogue: Catalogue) -> Tuple[CoordinateKey, ...]:
    """H37Rv coordinates for an MTBseq row, when they can be trusted.

    SNP rows give a clean ``(position, WT base, Var. base)`` triple. Insertions
    and deletions do not: MTBseq's encoding is not VCF-normalised — there is no
    anchor base and the columns describe the changed bases rather than the
    alleles — so an indel's coordinate is left out of the index rather than
    guessed at, and the row is still usable through its HGVS key.
    """
    var_type = header.value(row, MTBSEQ_COL_TYPE).upper()
    position = header.value(row, MTBSEQ_COL_POS_START)
    wt = header.value(row, MTBSEQ_COL_WT_BASE).upper()
    alt = header.value(row, MTBSEQ_COL_VAR_BASE).upper()
    if var_type != "SNP" or not position or not wt or not alt:
        if var_type and var_type != "SNP":
            catalogue._skip("indel coordinate not VCF-normalised; matched by name only")
        return ()
    if wt.strip("ACGTN") or alt.strip("ACGTN"):
        return ()
    try:
        pos = int(position)
    except ValueError:
        return ()
    return ((H37RV_ACCESSION, pos, wt, alt),)


# ---------------------------------------------------------------------------
# tbdb
# ---------------------------------------------------------------------------

def _tbdb_grade(confidence: str) -> str:
    """tbdb's confidence field as one of WHO's five canonical grade strings.

    tbdb re-publishes WHO v2's gradings with the numeric prefix removed, so
    ``Not assoc w R - Interim`` has to be mapped back onto
    ``4) Not assoc w R - Interim``. Prefixed spellings are tried too, through
    ``config.normalise_grade``, since tbdb's own upstream has used both.
    """
    text = _clean(confidence)
    if not text:
        return ""
    direct = normalise_grade(text)
    if direct:
        return direct
    return TBDB_CONFIDENCE_TO_GRADE.get(text.lower(), "")


def load_tbdb(path: PathLike) -> Catalogue:
    """Load tbdb's ``mutations.csv`` into the normal form.

    tbdb is TB-Profiler's library and mostly re-states WHO v2, which makes its
    independent value narrow but real: it carries rows WHO does not grade, and
    those are the §5.5 rule-3 path — reported as ``R (outside WHO catalogue)``,
    surfaced, and never presented as equivalent to a WHO Group 1 call.

    A row whose confidence field cannot be placed does not become susceptible.
    It becomes ``no-call`` and is counted, unless its ``type`` column asserts a
    resistance association on its own, in which case it is R without a grade.
    """
    resolved = Path(str(path)).expanduser()
    if not resolved.exists():
        raise MjolnirError(
            "tbdb mutations.csv not found at {0}.\n"
            "  fetch it with: {1}\n  (source: {2})"
            .format(resolved, FETCH_HINT, URL_TBDB))

    catalogue = Catalogue(
        name=CATALOGUE_TBDB, path=str(resolved), checksum=sha256sum(resolved),
        version="tbdb mutations.csv", licence=LICENCE_TBDB, citation=SRC_TBDB,
        url=URL_TBDB,
        note="tbdb largely re-states WHO v2 with the numeric grade prefixes "
             "removed, so agreement between the two is not independent evidence")

    observed_confidences: Dict[str, int] = {}
    with smart_open(resolved, "rt") as handle:
        reader = csv.reader(handle)
        try:
            header = _Header(next(reader))
        except StopIteration:
            raise MjolnirError("{0} is empty".format(resolved))
        missing = header.missing(("Gene", "Mutation", "drug"))
        if missing:
            raise MjolnirError(
                "{0} is missing the column(s) {1}; expected tbdb's header {2}.\n"
                "  re-fetch with: {3}".format(
                    resolved, ", ".join(repr(m) for m in missing),
                    ",".join(TBDB_COLUMNS), FETCH_HINT))

        row_number = 1
        for row in reader:
            row_number += 1
            gene = header.value(row, "Gene")
            mutation = header.value(row, "Mutation")
            drug_raw = header.value(row, "drug")
            if not gene or not mutation or not drug_raw:
                if any(_clean(cell) for cell in row):
                    catalogue._skip("row missing gene, mutation or drug")
                continue

            confidence = header.value(row, "confidence")
            row_type = header.value(row, "type").lower()
            if confidence:
                observed_confidences[confidence] = observed_confidences.get(confidence, 0) + 1
            grade = _tbdb_grade(confidence)
            if grade:
                call = call_for_grade(grade)
            elif row_type in TBDB_RESISTANCE_TYPES:
                call = CALL_R
            else:
                call = CALL_NO_CALL
                if confidence:
                    catalogue._skip("unrecognised confidence value")
                else:
                    catalogue._skip("row carries no confidence and no resistance type")

            hgvs = normalise_hgvs(mutation)
            original = header.value(row, "original_mutation")
            evidence = "; ".join(part for part in (
                "tbdb type: " + row_type if row_type else "",
                "tbdb source: " + header.value(row, "source")
                if header.value(row, "source") else "",
                "original mutation: " + original
                if original and original != mutation else "",
            ) if part)

            catalogue.entries.append(CatalogueEntry(
                catalogue=CATALOGUE_TBDB,
                drug=normalise_drug(drug_raw),
                gene=gene,
                hgvs=hgvs,
                call=call,
                grade=grade or confidence,
                comment=header.value(row, "comment"),
                effect="",
                evidence=evidence,
                coordinates=(),
                rule_only=is_rule_variant(hgvs),
                source_row=row_number,
            ))

    unplaced = [text for text in observed_confidences if not _tbdb_grade(text)]
    if unplaced:
        catalogue.discrepancies.append(
            "confidence value(s) tbdb uses that map to no WHO grade: {0}".format(
                ", ".join(sorted(repr(u) for u in unplaced))))
    catalogue.index()
    LOG.info("tbdb: %d (drug, variant) rows over %d variants, %d drugs",
             len(catalogue.entries), len(catalogue.variant_keys), len(catalogue.drugs))
    return catalogue


# ---------------------------------------------------------------------------
# Loading all three
# ---------------------------------------------------------------------------

def load_catalogues(db_dir: Optional[PathLike] = None,
                    who: Optional[PathLike] = None,
                    who_coordinates: Optional[PathLike] = None,
                    mtbseq: Optional[PathLike] = None,
                    tbdb: Optional[PathLike] = None,
                    strict: bool = True,
                    require_who: bool = True) -> Dict[str, Catalogue]:
    """Load whichever of the three catalogues are present.

    WHO is required by default because it is the anchor: §5.5 rule 2 makes the
    WHO grade the Mjolnir call wherever WHO grades the variant, and a consensus
    run without it is a different tool. The other two are optional, but their
    absence is logged with what it costs rather than passing unremarked — a run
    with only WHO cannot produce the ``R (outside WHO catalogue)`` result at
    all, and a reader needs to know that is why they did not see one.
    """
    root = Path(str(db_dir)).expanduser() if db_dir is not None else None
    loaded: Dict[str, Catalogue] = {}

    who_path = Path(str(who)) if who else (root / DEFAULT_WHO_XLSX if root else None)
    if who_path is not None and Path(str(who_path)).exists():
        coords = who_coordinates
        if coords is None and root is not None:
            candidate = root / DEFAULT_WHO_COORDINATES
            coords = candidate if candidate.exists() else None
        loaded[CATALOGUE_WHO] = load_who(who_path, coords, strict=strict)
    elif require_who:
        raise MjolnirError(
            "the WHO catalogue is required and was not found at {0}.\n"
            "  fetch it with: {1}\n"
            "  WHO is the anchor catalogue: where it grades a variant its grade "
            "is the Mjolnir call, and without it no graded resistance call can "
            "be made at all.".format(who_path, FETCH_HINT))

    mtbseq_path = Path(str(mtbseq)) if mtbseq else (
        root / DEFAULT_MTBSEQ_LIST if root else None)
    if mtbseq_path is not None and Path(str(mtbseq_path)).exists():
        loaded[CATALOGUE_MTBSEQ] = load_mtbseq(mtbseq_path)
    else:
        LOG.warning(
            "MTBseq resistance list not loaded (looked at %s); cross-catalogue "
            "agreement with MTBseq will be absent from this report, which is not "
            "the same as MTBseq agreeing", mtbseq_path)

    tbdb_path = Path(str(tbdb)) if tbdb else (root / DEFAULT_TBDB_CSV if root else None)
    if tbdb_path is not None and Path(str(tbdb_path)).exists():
        loaded[CATALOGUE_TBDB] = load_tbdb(tbdb_path)
    else:
        LOG.warning(
            "tbdb mutations.csv not loaded (looked at %s); variants graded by "
            "tbdb but not by WHO cannot be surfaced as 'R (outside WHO "
            "catalogue)' in this run", tbdb_path)

    return loaded


def database_versions(catalogues: Dict[str, Catalogue]) -> List[DatabaseVersion]:
    """One :class:`records.DatabaseVersion` per loaded catalogue, for the annex."""
    return [catalogues[name].database_version() for name in sorted(catalogues)]


def calls_for_variant(catalogues: Dict[str, Catalogue],
                      variant: Variant) -> List[CatalogueCall]:
    """Every catalogue's statement about one observed variant.

    Ordered with the anchor first, so that a reader scanning the annex sees the
    WHO grade before the two sources that cannot grade at all.
    """
    order = [CATALOGUE_WHO, CATALOGUE_MTBSEQ, CATALOGUE_TBDB]
    names = [n for n in order if n in catalogues] + \
        [n for n in sorted(catalogues) if n not in order]
    out: List[CatalogueCall] = []
    for name in names:
        out.extend(catalogues[name].calls_for(variant))
    return out


__all__ = [
    "Catalogue", "CatalogueEntry",
    "DEFAULT_MTBSEQ_LIST", "DEFAULT_TBDB_CSV", "DEFAULT_WHO_COORDINATES",
    "DEFAULT_WHO_XLSX", "FETCH_HINT",
    "LICENCE_MTBSEQ", "LICENCE_TBDB", "LICENCE_WHO",
    "TBDB_CONFIDENCE_TO_GRADE", "WHO_COORD_COLUMNS", "WHO_REQUIRED_COLUMNS",
    "WHO_TXT_REFUSAL",
    "calls_for_variant", "database_versions", "load_catalogues", "load_mtbseq",
    "load_tbdb", "load_who", "read_who_coordinates_file",
    "refuse_who_text_master",
]
